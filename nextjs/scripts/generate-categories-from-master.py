#!/usr/bin/env python3
"""Generate marketplace category tree + listing metadata from Excel master file."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "eseller_angilal_master.xlsx"
OUT_CAT = ROOT / "src" / "lib" / "generated" / "categoryMaster.ts"
OUT_META = ROOT / "src" / "lib" / "generated" / "categoryAttributes.ts"

# Excel root code → stable system key + UI meta (preserve existing keys)
ROOT_META: dict[str, dict] = {
    "01": {
        "key": "jobs",
        "label": "Ажлын зар",
        "shortLabel": "Ажил",
        "emoji": "💼",
        "icon": "BriefcaseBusiness",
        "color": "#4F46E5",
        "section": "listing",
        "entityTypes": ["STORE", "SERVICE", "USER", "COMPANY"],
        "aliases": ["job", "jobs", "work", "career", "hiring", "part-time", "ажил"],
    },
    "02": {
        "key": "women",
        "label": "Эмэгтэй",
        "emoji": "👗",
        "icon": "Venus",
        "color": "#DB2777",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["fashion", "female", "women-fashion", "эмэгтэй"],
    },
    "03": {
        "key": "men",
        "label": "Эрэгтэй",
        "emoji": "👔",
        "icon": "Mars",
        "color": "#2563EB",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["men-fashion", "эрэгтэй"],
    },
    "04": {
        "key": "beauty-health",
        "label": "Гоо сайхан",
        "emoji": "💄",
        "icon": "Sparkles",
        "color": "#C026D3",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["beauty", "salon", "гоо"],
    },
    "05": {
        "key": "home-decor",
        "label": "Гэр декор",
        "emoji": "🏡",
        "icon": "Home",
        "color": "#059669",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["home", "decor", "interior", "гэр"],
    },
    "06": {
        "key": "real-estate",
        "label": "Үл хөдлөх",
        "emoji": "🏢",
        "icon": "Building2",
        "color": "#0F766E",
        "section": "listing",
        "entityTypes": ["REAL_ESTATE", "STORE", "COMPANY"],
        "aliases": ["property", "apartment", "house", "үл-хөдлөх", "ul-hudluh"],
    },
    "07": {
        "key": "new-buildings",
        "label": "Шинэ төсөл",
        "emoji": "🏗️",
        "icon": "Construction",
        "color": "#B45309",
        "section": "listing",
        "entityTypes": ["REAL_ESTATE", "COMPANY"],
        "aliases": ["new-building", "project", "шинэ-төсөл"],
    },
    "08": {
        "key": "vehicles",
        "label": "Машин",
        "emoji": "🚗",
        "icon": "Car",
        "color": "#DC2626",
        "section": "listing",
        "entityTypes": ["AUTO", "STORE"],
        "aliases": ["auto", "car", "cars", "vehicle", "машин"],
    },
    "09": {
        "key": "jewelry",
        "label": "Гоёл",
        "emoji": "💍",
        "icon": "Gem",
        "color": "#A855F7",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["jewellery", "accessories-jewelry", "гоёл"],
    },
    "10": {
        "key": "kids",
        "label": "Хүүхдийн",
        "emoji": "🧸",
        "icon": "Baby",
        "color": "#F59E0B",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["children", "baby", "хүүхэд", "kids-toys"],
    },
    "11": {
        "key": "adult",
        "label": "18+",
        "emoji": "🔞",
        "icon": "Shield",
        "color": "#7C3AED",
        "section": "product",
        "entityTypes": ["STORE"],
        "aliases": ["18plus", "adult"],
    },
    "12": {
        "key": "health",
        "label": "Эрүүл мэнд",
        "emoji": "💊",
        "icon": "HeartPulse",
        "color": "#16A34A",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["medicine", "pharmacy", "эрүүл", "health-vitamins"],
    },
    "13": {
        "key": "phones",
        "label": "Утас",
        "emoji": "📱",
        "icon": "Smartphone",
        "color": "#0284C7",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["phone", "mobile", "smartphone", "утас"],
    },
    "14": {
        "key": "technology",
        "label": "Технологи",
        "emoji": "💻",
        "icon": "Laptop",
        "color": "#0891B2",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER", "DIGITAL"],
        "aliases": ["tech", "electronics", "компьютер", "технологи"],
    },
    "15": {
        "key": "gifts",
        "label": "Бэлэг",
        "emoji": "🎁",
        "icon": "Gift",
        "color": "#E11D48",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["gift", "бэлэг", "gifts-hobby"],
    },
    "16": {
        "key": "furniture",
        "label": "Тавилга",
        "emoji": "🛋️",
        "icon": "Armchair",
        "color": "#92400E",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["тавилга"],
    },
    "17": {
        "key": "appliances",
        "label": "Цахилгаан",
        "emoji": "🔌",
        "icon": "Plug",
        "color": "#EA580C",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["appliance", "цахилгаан"],
    },
    "18": {
        "key": "auto-parts",
        "label": "Авто сэлбэг",
        "emoji": "🛠️",
        "icon": "Wrench",
        "color": "#B91C1C",
        "section": "product",
        "entityTypes": ["STORE", "AUTO"],
        "aliases": ["parts", "spare-parts", "сэлбэг"],
    },
    "19": {
        "key": "books",
        "label": "Ном",
        "emoji": "📚",
        "icon": "BookOpen",
        "color": "#1D4ED8",
        "section": "product",
        "entityTypes": ["STORE", "DIGITAL"],
        "aliases": ["book", "ном", "books-stationery"],
    },
    "20": {
        "key": "construction",
        "label": "Барилга",
        "emoji": "🧱",
        "icon": "Construction",
        "color": "#78716C",
        "section": "product",
        "entityTypes": ["STORE", "COMPANY"],
        "aliases": ["building-materials", "барилга", "construction-tools"],
    },
    "21": {
        "key": "travel",
        "label": "Аялал",
        "emoji": "✈️",
        "icon": "TentTree",
        "color": "#0EA5E9",
        "section": "product",
        "entityTypes": ["STORE", "SERVICE"],
        "aliases": ["tourism", "аялал"],
    },
    "22": {
        "key": "sports",
        "label": "Спорт",
        "emoji": "⚽",
        "icon": "Dumbbell",
        "color": "#16A34A",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["sport", "спорт"],
    },
    "23": {
        "key": "food",
        "label": "Хүнс",
        "emoji": "🍲",
        "icon": "UtensilsCrossed",
        "color": "#F97316",
        "section": "product",
        "entityTypes": ["STORE", "PRE_ORDER"],
        "aliases": ["grocery", "хүнс", "food-beverage"],
    },
    "24": {
        "key": "esports",
        "label": "E-Спорт",
        "emoji": "🎮",
        "icon": "Gamepad2",
        "color": "#7C3AED",
        "section": "product",
        "entityTypes": ["STORE", "DIGITAL"],
        "aliases": ["gaming", "e-sport"],
    },
    "25": {
        "key": "pets-plants",
        "label": "Амьтан ургамал",
        "emoji": "🐾",
        "icon": "Dog",
        "color": "#65A30D",
        "section": "product",
        "entityTypes": ["STORE"],
        "aliases": ["pets", "plants", "амьтан"],
    },
    "26": {
        "key": "digital",
        "label": "Дижитал",
        "emoji": "💾",
        "icon": "Monitor",
        "color": "#6366F1",
        "section": "product",
        "entityTypes": ["STORE", "DIGITAL"],
        "aliases": ["digital-goods", "дижитал", "digital-goods"],
    },
    "27": {
        "key": "education-training",
        "label": "Сургалт",
        "emoji": "🎓",
        "icon": "GraduationCap",
        "color": "#2563EB",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["tutoring", "боловсрол", "сургалт"],
    },
    "28": {
        "key": "beauty-services",
        "label": "Гоо үйлчилгээ",
        "emoji": "💅",
        "icon": "Scissors",
        "color": "#DB2777",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["salon", "beauty_service", "haircut"],
    },
    "29": {
        "key": "tech-it-services",
        "label": "IT үйлчилгээ",
        "emoji": "🧑‍💻",
        "icon": "Laptop",
        "color": "#0891B2",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["web_dev", "design-it", "it"],
    },
    "30": {
        "key": "professional-consulting",
        "label": "Зөвлөх",
        "emoji": "💼",
        "icon": "BriefcaseBusiness",
        "color": "#4F46E5",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["consulting", "зөвлөх"],
    },
    "31": {
        "key": "auto-services",
        "label": "Авто үйлчилгээ",
        "emoji": "🚙",
        "icon": "Car",
        "color": "#DC2626",
        "section": "service",
        "entityTypes": ["SERVICE", "AUTO"],
        "aliases": ["car-service"],
    },
    "32": {
        "key": "repair-services",
        "label": "Засвар",
        "emoji": "🔧",
        "icon": "Wrench",
        "color": "#EA580C",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["repair", "засвар"],
    },
    "33": {
        "key": "printing-services",
        "label": "Хэвлэл",
        "emoji": "🖨️",
        "icon": "Printer",
        "color": "#475569",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["printing", "хэвлэл"],
    },
    "34": {
        "key": "manufacturing-custom",
        "label": "Захиалгат",
        "emoji": "🏭",
        "icon": "Factory",
        "color": "#7C2D12",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["factory", "custom"],
    },
    "35": {
        "key": "photo-video",
        "label": "Зураг авалт",
        "emoji": "📷",
        "icon": "Camera",
        "color": "#9333EA",
        "section": "service",
        "entityTypes": ["SERVICE"],
        "aliases": ["photo", "video"],
    },
    "36": {
        "key": "design-creative",
        "label": "Дизайн",
        "emoji": "🎨",
        "icon": "Palette",
        "color": "#C026D3",
        "section": "service",
        "entityTypes": ["SERVICE", "DIGITAL"],
        "aliases": ["design", "дизайн"],
    },
}


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[\s/,&+]+", "-", value)
    value = re.sub(r"[^\w\-а-яөүёА-ЯӨҮЁ0-9]+", "", value, flags=re.UNICODE)
    value = re.sub(r"-{2,}", "-", value).strip("-")
    return value or "item"


def ts_str(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def parse_options(raw: str | None) -> list[str] | None:
    if not raw or not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text or text in {"—", "-", "Дэд ангиллын жагсаалт", "Стандарт жагсаалт"}:
        return None
    # Split common separators
    if "/" in text and "→" not in text and len(text) < 200:
        parts = [p.strip() for p in re.split(r"[/]", text) if p.strip()]
        # filter noisy
        if 2 <= len(parts) <= 20 and all(len(p) < 40 for p in parts):
            cleaned = []
            for p in parts:
                p = re.sub(r"\.\.\.$", "", p).strip()
                p = re.sub(r"нэгдсэн жагсаалт.*", "", p, flags=re.I).strip()
                if p and "жагсаалт" not in p.lower() and "хамаар" not in p.lower():
                    cleaned.append(p)
            return cleaned or None
    if " / " in text:
        parts = [p.strip() for p in text.split(" / ") if p.strip()]
        if 2 <= len(parts) <= 20:
            return parts
    return None


def map_type(raw: str | None) -> str:
    t = (raw or "Текст").strip().lower()
    if "сонголт" in t or "тийм/үгүй" in t:
        if "тийм" in t or "үгүй" in t:
            return "boolean"
        return "select"
    if "тоо" in t:
        return "number"
    if "файл" in t:
        return "text"
    return "text"


def field_key(label: str) -> str:
    mapping = {
        "Салбар": "sector",
        "Цалин (муж)": "salaryRange",
        "Ажлын цаг": "jobType",
        "Шаардлагатай туршлага": "experience",
        "Боловсрол": "education",
        "Компанийн нэр": "employerName",
        "Хэмжээ": "size",
        "Өнгө": "color",
        "Брэнд": "brand",
        "Материал": "material",
        "Улирал": "season",
        "Хэмжээ, багтаамж": "sizeVolume",
        "Хугацаа дуусах огноо": "expiryDate",
        "Гарал улс": "originCountry",
        "Арьсны төрөл": "skinType",
        "Хэмжээ (см)": "dimensionsCm",
        "Үйлдвэрлэгч": "brand",
        "Загвар": "model",
        "Үйлдвэрлэсэн он": "year",
        "Орж ирсэн он": "importYear",
        "Гүйлт (км)": "mileage",
        "Хөдөлгүүр (л)": "engine",
        "Түлш": "fuelType",
        "Хурдны хайрцаг": "transmission",
        "Хөтлөгч": "drivetrain",
        "Улсын дугаартай эсэх": "hasPlate",
        "Лизингтэй эсэх": "leasing",
        "Албан тушаал": "positionTitle",
        "Ажил олгогч": "employerName",
        "Ажиллах орчин": "workplaceType",
        "Хуваарь": "schedule",
        "Талбай, м²": "sqm",
        "Өрөө": "rooms",
        "Төслийн төлөв": "projectStatus",
        "1м² үнэ, ₮": "pricePerSqm",
        "Багтаамж": "storage",
        "Battery health, %": "batteryHealth",
        "Төлөв": "condition",
        "Баталгаа": "warranty",
    }
    if label in mapping:
        return mapping[label]
    # fallback slug latin-ish
    s = slugify(label)
    s = re.sub(r"[^a-z0-9\-]", "", s) or "field"
    s = s.replace("-", "_")
    # camelCase-ish
    parts = [p for p in s.split("_") if p]
    if not parts:
        return "field"
    return parts[0] + "".join(p.title() for p in parts[1:])


def required_flag(raw: str | None) -> bool:
    t = (raw or "").strip().lower()
    return t.startswith("заавал")


def load_tree(wb):
    ws = wb["Ангиллын мод"]
    nodes = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        code = str(r[0]).strip()
        name = str(r[1]).strip()
        level = int(r[2])
        nodes.append({"code": code, "name": name, "level": level})
    return nodes


def build_branches(nodes, root_code: str):
    """Build MarketplaceCategoryBranch tree under a root code."""
    children_by_parent: dict[str, list[dict]] = defaultdict(list)
    for n in nodes:
        code = n["code"]
        if code == root_code or not code.startswith(root_code):
            continue
        # parent code = strip last segment
        if "-" not in code:
            continue
        parent = code.rsplit("-", 1)[0]
        children_by_parent[parent].append(n)

    def to_branch(node: dict):
        kids = children_by_parent.get(node["code"], [])
        if not kids:
            return node["name"]
        return {
            "name": node["name"],
            "aliases": [node["code"], slugify(node["name"])],
            "children": [to_branch(k) for k in kids],
        }

    top = children_by_parent.get(root_code, [])
    return [to_branch(n) for n in top]


def load_attributes(wb):
    ws = wb["Ангиллын атрибут"]
    by_code: dict[str, list[dict]] = defaultdict(list)
    cur = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        if r[0]:
            cur = str(r[0]).strip()
        if not cur:
            continue
        label = str(r[2]).strip() if r[2] else ""
        if not label:
            continue
        by_code[cur].append(
            {
                "label": label,
                "typeRaw": r[3],
                "requiredRaw": r[4],
                "optionsRaw": r[5],
                "rule": r[6],
                "usage": r[7],
            }
        )
    return by_code


def emit_branch_ts(branch, indent: int) -> str:
    pad = "  " * indent
    if isinstance(branch, str):
        return f"{pad}{ts_str(branch)}"
    lines = [f"{pad}{{"]
    lines.append(f"{pad}  name: {ts_str(branch['name'])},")
    if branch.get("aliases"):
        alias_list = ", ".join(ts_str(a) for a in branch["aliases"])
        lines.append(f"{pad}  aliases: [{alias_list}],")
    if branch.get("children"):
        lines.append(f"{pad}  children: [")
        for i, child in enumerate(branch["children"]):
            comma = "," if i < len(branch["children"]) - 1 else ""
            lines.append(emit_branch_ts(child, indent + 2) + comma)
        lines.append(f"{pad}  ],")
    lines.append(f"{pad}}}")
    return "\n".join(lines)


def emit_category_ts(meta: dict, branches: list, indent: int = 1) -> str:
    pad = "  " * indent
    lines = [f"{pad}{{"]
    lines.append(f"{pad}  key: {ts_str(meta['key'])},")
    lines.append(f"{pad}  label: {ts_str(meta['label'])},")
    if meta.get("shortLabel"):
        lines.append(f"{pad}  shortLabel: {ts_str(meta['shortLabel'])},")
    lines.append(f"{pad}  emoji: {ts_str(meta['emoji'])},")
    lines.append(f"{pad}  icon: {ts_str(meta['icon'])},")
    lines.append(f"{pad}  color: {ts_str(meta['color'])},")
    lines.append(f"{pad}  section: {ts_str(meta['section'])},")
    et = ", ".join(ts_str(x) for x in meta["entityTypes"])
    lines.append(f"{pad}  entityTypes: [{et}],")
    if meta.get("aliases"):
        al = ", ".join(ts_str(a) for a in meta["aliases"])
        lines.append(f"{pad}  aliases: [{al}],")
    lines.append(f"{pad}  subcategories: [")
    for i, b in enumerate(branches):
        comma = "," if i < len(branches) - 1 else ""
        lines.append(emit_branch_ts(b, indent + 2) + comma)
    lines.append(f"{pad}  ],")
    lines.append(f"{pad}}}")
    return "\n".join(lines)


def emit_field_ts(field: dict, indent: int = 1) -> str:
    pad = "  " * indent
    parts = [
        f"key: {ts_str(field['key'])}",
        f"label: {ts_str(field['label'])}",
        f"type: {ts_str(field['type'])}",
    ]
    if field.get("required"):
        parts.append("required: true")
    if field.get("options"):
        opts = ", ".join(ts_str(o) for o in field["options"])
        parts.append(f"options: [{opts}]")
    if field.get("placeholder"):
        parts.append(f"placeholder: {ts_str(field['placeholder'])}")
    if field.get("hint"):
        parts.append(f"hint: {ts_str(field['hint'])}")
    return f"{pad}{{ {', '.join(parts)} }}"


def main():
    if not XLSX.exists():
        raise SystemExit(f"Missing Excel: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    nodes = load_tree(wb)
    attrs = load_attributes(wb)

    product_cats = []
    service_cats = []
    for code, meta in sorted(ROOT_META.items(), key=lambda x: x[0]):
        branches = build_branches(nodes, code)
        block = {"meta": meta, "branches": branches, "code": code}
        if meta["section"] == "service":
            service_cats.append(block)
        else:
            product_cats.append(block)

    # categoryMaster.ts
    lines = [
        "/** AUTO-GENERATED from data/eseller_angilal_master.xlsx — do not edit by hand. */",
        "/** Run: python scripts/generate-categories-from-master.py */",
        "",
        "import type { MarketplaceCategory } from '../marketplaceCategories';",
        "",
        "export const GENERATED_PRODUCT_MARKETPLACE_CATEGORIES: MarketplaceCategory[] = [",
    ]
    for i, block in enumerate(product_cats):
        comma = "," if i < len(product_cats) - 1 else ""
        lines.append(emit_category_ts(block["meta"], block["branches"], 1) + comma)
    lines.append("];")
    lines.append("")
    lines.append("export const GENERATED_SERVICE_MARKETPLACE_CATEGORIES: MarketplaceCategory[] = [")
    for i, block in enumerate(service_cats):
        comma = "," if i < len(service_cats) - 1 else ""
        lines.append(emit_category_ts(block["meta"], block["branches"], 1) + comma)
    lines.append("];")
    lines.append("")
    lines.append("export const GENERATED_MARKETPLACE_CATEGORIES: MarketplaceCategory[] = [")
    lines.append("  ...GENERATED_PRODUCT_MARKETPLACE_CATEGORIES,")
    lines.append("  ...GENERATED_SERVICE_MARKETPLACE_CATEGORIES,")
    lines.append("];")
    lines.append("")

    OUT_CAT.parent.mkdir(parents=True, exist_ok=True)
    OUT_CAT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # attributes
    meta_lines = [
        "/** AUTO-GENERATED from data/eseller_angilal_master.xlsx — do not edit by hand. */",
        "/** Run: python scripts/generate-categories-from-master.py */",
        "",
        "import type { ListingMetadataField } from '../listingMetadata';",
        "",
        "export const GENERATED_CATEGORY_ATTRIBUTES: Record<string, ListingMetadataField[]> = {",
    ]

    used_keys_global: set[str] = set()
    for code, meta in sorted(ROOT_META.items(), key=lambda x: x[0]):
        fields_raw = attrs.get(code, [])
        fields = []
        seen = set()
        for fr in fields_raw:
            key = field_key(fr["label"])
            # uniquify within category
            base = key
            n = 2
            while key in seen:
                key = f"{base}{n}"
                n += 1
            seen.add(key)
            ftype = map_type(str(fr["typeRaw"] or ""))
            options = parse_options(str(fr["optionsRaw"]) if fr["optionsRaw"] else None)
            if ftype == "boolean":
                options = None
            if ftype == "select" and not options:
                # keep select only if we have options; else text
                ftype = "text"
            field = {
                "key": key,
                "label": fr["label"],
                "type": ftype,
                "required": required_flag(str(fr["requiredRaw"] or "")),
            }
            if options:
                field["options"] = options
            if fr.get("rule"):
                field["hint"] = str(fr["rule"])[:120]
            fields.append(field)
            used_keys_global.add(key)

        meta_lines.append(f"  {ts_str(meta['key'])}: [")
        for i, f in enumerate(fields):
            comma = "," if i < len(fields) - 1 else ""
            meta_lines.append(emit_field_ts(f, 2) + comma)
        meta_lines.append("  ],")

    meta_lines.append("};")
    meta_lines.append("")
    meta_lines.append("export function generatedFieldsForCategoryKey(key: string): ListingMetadataField[] {")
    meta_lines.append("  return GENERATED_CATEGORY_ATTRIBUTES[key] || [];")
    meta_lines.append("}")
    meta_lines.append("")

    OUT_META.write_text("\n".join(meta_lines) + "\n", encoding="utf-8")

    print(f"Wrote {OUT_CAT.relative_to(ROOT)}")
    print(f"  product roots: {len(product_cats)}")
    print(f"  service roots: {len(service_cats)}")
    print(f"Wrote {OUT_META.relative_to(ROOT)}")
    print(f"  attribute keys: {len(used_keys_global)}")
    # node counts
    print(f"  excel nodes: {len(nodes)}")


if __name__ == "__main__":
    main()
